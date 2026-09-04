#!/usr/bin/env python3
"""
발표용 시연 엑셀 세트 — 끌어다 놓는 순서대로 번호를 붙였다.

왜 파일로 만드는가
    시연 중에 프롬프트를 복사·붙여넣기 하면 오타가 나고, 오타 하나로 판정이 달라진다.
    파일을 끌어다 놓으면 같은 입력이 매번 같은 판정을 낸다.

무엇을 보이는가 (파일 순서 = 발표 순서)
    1   통과        게이트웨이는 전부 막는 물건이 아니다
    2a  차단        자격증명·주민번호는 나가기 전에 멈춘다
    2b  재전송      고쳐서 다시 보내면 통과한다
    3   부서 차이   같은 문장이 영업팀에서는 검토, 개발팀에서는 통과
    4   엠바고 차단 아직 때가 아니라서 막는다
    5   엠바고 해제 해제일이 지나면 같은 종류의 문장이 통과한다
    6   오탐        전화번호 형식의 부품 번호가 가려진다 → 해제 검토로 사람이 푼다
    7   출력 검사   나간 뒤 돌아온 답변이 코드를 되돌려주면 유출로 잡는다

    예비  마스킹    고객 명단 규칙. 한계만 드러내고 그 자리에서 풀 방법이 없어
                    본 순서에서 뺐다. 명단·동명이인 질문이 나오면 그때 꺼낸다.

데이터는 전부 합성이다. 실제 고객·연락처·계약·티켓이 아니다.
담당자는 기획서 10.1 원칙에 따라 `이OO` 형식을 쓴다.
고객명은 시드 명단의 이름을 쓴다 — 명단 규칙이 걸리는 것을 보여야 하기 때문이다.

각 파일이 어느 계정에서 어떤 판정을 내야 하는지는 EXPECTED에 적어 두었고,
`--verify` 를 주면 배포 서버에 실제로 넣어 대조한다.

실행
    /Users/jeonghangyeol/workspace/.venv/bin/python3 make_demo_files.py
    /Users/jeonghangyeol/workspace/.venv/bin/python3 make_demo_files.py --verify
"""

import json
import sys
import urllib.request
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).parent
SERVER = "http://15.164.215.132/api/v1"

# 계정 — 시드 사용자. 부서마다 적용 정책이 다르다.
USERS = {"이OO": 1, "김OO": 2, "정OO": 3, "박OO": 4, "한OO": 5}
DEPT = {1: "개발팀", 2: "영업팀", 3: "인사팀", 4: "정보보안팀", 5: "홍보팀"}

# 붙일 때 함께 적을 질문. 파일마다 다르다 — 표만 던지는 것은 실제 사용 모습이 아니다.
# 질문 자체가 규칙에 걸리면 시연 의도가 흐려지므로 키워드를 피해서 썼다.
PROMPTS = {}


def sheet(wb, title, rows, widths=None, wrap_col=None):
    ws = wb.create_sheet(title) if wb.sheetnames != ["Sheet"] else wb.active
    ws.title = title
    for row in rows:
        ws.append(row)
    head = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="2F5D8A")
    for cell in ws[1]:
        cell.font, cell.fill = head, fill
        cell.alignment = Alignment(horizontal="center")
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        for i, col in enumerate(ws.columns, 1):
            width = max(len(str(c.value or "")) for c in col) + 4
            ws.column_dimensions[get_column_letter(i)].width = max(10, width)
    if wrap_col:
        for row in ws.iter_rows(min_row=2, min_col=wrap_col, max_col=wrap_col):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    return ws


def build(name, title, rows, prompt, widths=None, wrap_col=None):
    wb = Workbook()
    sheet(wb, title, rows, widths, wrap_col)
    path = OUT_DIR / name
    wb.save(path)
    PROMPTS[name] = prompt
    return path


# ── 01. 통과 ───────────────────────────────────────────────────────────────
# 걸릴 것이 하나도 없다. 기본 경로가 통제가 아니라 통과라는 것을 먼저 보인다.
F01 = build(
    "1_통과_스프린트회고.xlsx", "스프린트 회고",
    [
        ["구분", "항목", "담당", "비고"],
        ["잘된 것", "배포 파이프라인 자동화로 릴리스 시간 40분 단축", "이OO", "다음 분기 유지"],
        ["잘된 것", "테스트 커버리지 62% → 78%", "박OO", ""],
        ["아쉬운 것", "코드 리뷰 대기 시간이 평균 1.5일", "이OO", "리뷰어 2인 배정"],
        ["아쉬운 것", "스테이징 환경 재현이 어려움", "박OO", "컨테이너 이미지 통일"],
        ["다음 할 일", "장애 대응 런북 정리", "이OO", "10월 첫 주"],
    ],
    "이번 스프린트 회고를 정리해서 다음 분기 개선 과제 세 가지만 뽑아줘.",
)

# ── 02. 마스킹 + 규칙의 한계 ────────────────────────────────────────────────
# 김서준·박예린은 시드 명단에 있고 최도윤·정하윤은 없다. 이름은 명단에 있는 것만
# 잡히고, 형식이 있는 연락처·이메일은 명단과 무관하게 전부 잡힌다.
# 계약금액은 규칙이 없어 그대로 나간다 — 무엇을 민감하다고 볼지는 보안팀이 정한다.
F02 = build(
    "예비_마스킹_고객명단.xlsx", "3분기 VIP",
    [
        ["고객명", "담당자", "연락처", "이메일", "계약금액", "비고"],
        ["김서준", "이OO", "010-2841-7734", "seojun.kim@sy-tech.co.kr", 120000000, "재계약 협의 중"],
        ["박예린", "김OO", "010-9925-1180", "yerin.park@kkenergy.kr", 84500000, "견적 재요청"],
        ["최도윤", "이OO", "010-3317-4402", "doyun.choi@daehan-log.co.kr", 61200000, "신규 문의"],
        ["정하윤", "김OO", "010-4408-2251", "hayun.jung@nseng.co.kr", 45800000, "보류"],
    ],
    "이 표에서 계약금액 상위 고객만 추려서 재계약 우선순위를 정리해줘.",
)

# ── 03a. 차단 ──────────────────────────────────────────────────────────────
# DB 접속 문자열(자격증명)과 주민번호. 둘 다 있으면 가장 강한 조치인 차단이 이긴다.
F03A = build(
    "2a_차단_접속정보.xlsx", "이관 체크리스트",
    [
        ["항목", "내용", "담당"],
        ["대상 시스템", "정산 배치 v2", "이OO"],
        ["접속 문자열", "postgres://admin:p%40ss@10.0.3.21/prod", "이OO"],
        ["점검 계정", "담당자 주민번호 900101-1234567 기준으로 권한 확인", "이OO"],
        ["일정", "10월 2주차 이관", "박OO"],
    ],
    "이 이관 체크리스트에서 빠진 점검 항목이 있는지 봐줘.",
    widths=[16, 62, 10],
)

# ── 03b. 고쳐서 재전송 ──────────────────────────────────────────────────────
# 같은 일을 하려는 같은 표다. 자격증명을 환경변수 이름으로 바꾸고 주민번호를 뺐다.
# 차단은 끝이 아니라 고쳐 쓰라는 안내라는 것을 보인다.
F03B = build(
    "2b_재전송_통과.xlsx", "이관 체크리스트",
    [
        ["항목", "내용", "담당"],
        ["대상 시스템", "정산 배치 v2", "이OO"],
        ["접속 문자열", "환경변수 DB_URL 사용 (값은 시크릿 매니저)", "이OO"],
        ["점검 계정", "사번 기준으로 권한 확인", "이OO"],
        ["일정", "10월 2주차 이관", "박OO"],
    ],
    "이 이관 체크리스트에서 빠진 점검 항목이 있는지 봐줘.",
    widths=[16, 62, 10],
)

# ── 04. 같은 문장, 다른 부서 ────────────────────────────────────────────────
# CONF-CLIENT-01(고객사명·차세대)은 영업팀·인사팀에만 적용된다. 개발팀에는 없다.
# 같은 파일을 두 계정으로 붙여 판정이 갈리는 것을 보인다.
F04 = build(
    "3_부서차이_고객사미팅.xlsx", "미팅 메모",
    [
        ["일시", "안건", "내용", "작성"],
        ["09-02", "요구사항 정리", "A사 차세대 시스템 이관 범위 협의", "김OO"],
        ["09-02", "일정", "설계 검토 2주, 이관 리허설 1주", "김OO"],
        ["09-03", "위험", "레거시 배치 의존성 확인 필요", "김OO"],
        ["09-03", "다음 회의", "산정 결과 공유 후 재논의", "김OO"],
    ],
    "이 미팅 메모를 바탕으로 다음 회의 안건을 정리해줘.",
    widths=[10, 18, 52, 10],
)

# ── 05. 엠바고 차단 ────────────────────────────────────────────────────────
# 해제일 2026-09-20. 발표 당일 기준 아직 유효하다.
# 고객사 키워드를 피해 엠바고 규칙 하나만 걸리게 했다.
F05 = build(
    "4_엠바고_차단_노바.xlsx", "런칭 준비",
    [
        ["구분", "항목", "담당", "상태"],
        ["기능", "SKALA NOVA 온보딩 플로우 1차", "이OO", "진행"],
        ["기능", "노바 추천 위젯 지표 확정", "박OO", "진행"],
        ["일정", "SKALA NOVA 정식 오픈 준비", "이OO", "대기"],
        ["홍보", "노바 랜딩 페이지 문구 검토", "한OO", "대기"],
    ],
    "이 준비 항목들을 우선순위대로 정리해줘.",
)

# ── 06. 엠바고 해제 ────────────────────────────────────────────────────────
# 해제일 2026-09-04. 발표 당일에 풀린다. 같은 종류의 문장이 통과한다.
F06 = build(
    "5_엠바고_해제_아틀라스.xlsx", "개선 백로그",
    [
        ["구분", "항목", "담당", "상태"],
        ["기능", "ATLAS 대시보드 위젯 추가", "이OO", "진행"],
        ["기능", "아틀라스 리포트 내보내기", "박OO", "진행"],
        ["성능", "ATLAS 필터 응답 개선 p95 1.2s → 400ms", "이OO", "완료"],
        ["운영", "아틀라스 권한 그룹 정리", "박OO", "완료"],
    ],
    "이 개선 항목들을 우선순위대로 정리해줘.",
)

# ── 07. 오탐 → 사람이 푼다 ──────────────────────────────────────────────────
# 부품 번호가 전화번호 형식이다. 규칙은 형식만 보므로 가린다.
# 마스킹 해제 검토(D25)로 사람이 원문과 마스킹본을 비교해 확정하는 자리다.
F07 = build(
    "6_오탐_부품번호.xlsx", "자재 발주",
    [
        ["부품 코드", "품명", "수량", "단가", "납기"],
        ["010-2841-7734", "베어링 하우징 A형", 120, 18500, "10-08"],
        ["010-9925-1180", "구동축 커플러", 80, 42000, "10-11"],
        ["031-2211-9080", "감속기 커버", 45, 66000, "10-15"],
        ["070-8842-3310", "센서 브래킷", 200, 7300, "10-18"],
    ],
    "이 발주 목록에서 납기가 촉박한 순서로 정리해줘.",
)

# ── 08. 출력 검사 ──────────────────────────────────────────────────────────
# 재현 코드를 한 셀에 넣는다. 추출하면 한 줄이 되고, 모델이 그 줄을 고쳐서
# 통째로 되돌려준다. 사내 코드가 외부를 한 바퀴 돌아 그대로 나온 장면이다.
# 되돌아온 조각이 40자를 넘어야 유출 의심이 걸린다 — 아래에서 길이를 잰다.
REPRO_CODE = (
    "RetryPolicy policy = config.getRetryPolicy();\n"
    "int max = policy.getMaxAttempts();\n"
    'client.connect("10.0.3.21", 5432);'
)
F08 = build(
    "7_출력검사_리포트_최종본.xlsx", "장애 리포트",
    [
        ["항목", "내용"],
        ["티켓 번호", "INC-2026-0903-17"],
        ["발생 시각", "2026-09-03 14:22"],
        ["발생 모듈", "결제 재시도 (payment-retry)"],
        ["심각도", "P2"],
        ["증상", "결제 재시도에서 NPE 나는데 봐줘."],
        ["재현 코드", REPRO_CODE],
        ["DB 호스트", "10.0.3.21:5432"],
        ["담당자", "이OO"],
        ["담당자 연락처", "010-3317-4402"],
        ["조치 상태", "원인 분석 중"],
        # 유출 검사의 두 번째 그물. 모델이 코드를 되돌려주면 코드 되돌림 검사가 잡고,
        # 표를 요약해 버리면 이 줄을 옮겨 적으면서 문맥 검사가 잡는다.
        # 한 겹에만 기대면 모델이 어떻게 답하느냐에 시연이 좌우된다.
        ["공유 범위", "대외비 — 사내 개발팀 한정. 외부 공유 금지"],
    ],
    # 질문이 결과를 가른다. "원인 좀 봐줘"로 물으면 모델이 표를 요약해 버리고,
    # 그러면 되돌아온 조각이 문턱(40자)을 못 넘어 유출 검사가 걸리지 않는다.
    # **고친 코드를 달라고 해야** 모델이 코드를 통째로 되돌려준다 — 그 장면이 이 파일의 목적이다.
    "재현 코드는 그대로 두고 null 체크 한 줄만 추가해서 전체 코드를 다시 써줘.",
    widths=[16, 62], wrap_col=2,
)

# 파일 → (계정, 기대 판정, 한 줄 설명)
EXPECTED = [
    (F01, [("이OO", "ALLOW", "걸릴 것이 없다. 원문 그대로 나간다")]),
    (F02, [("김OO", "MASK", "명단에 있는 이름·연락처·이메일만 가려진다. 최도윤·정하윤·계약금액은 그대로")]),
    (F03A, [("이OO", "BLOCK", "자격증명과 주민번호. 나가기 전에 멈춘다")]),
    (F03B, [("이OO", "ALLOW", "고쳐서 다시 보내면 통과한다")]),
    (F04, [("김OO", "PENDING", "영업팀에는 고객사 정책이 적용된다 → 검토 대기"),
           ("이OO", "ALLOW", "개발팀에는 그 정책이 없다 → 같은 문장이 통과")]),
    (F05, [("이OO", "BLOCK", "해제일 2026-09-20. 아직 못 연다"),
           ("한OO", "ALLOW", "홍보팀은 엠바고를 정한 쪽이라 적용 대상이 아니다")]),
    (F06, [("이OO", "ALLOW", "해제일 2026-09-04. 오늘부터 열린다")]),
    (F07, [("김OO", "MASK", "부품 번호가 전화번호 형식이라 가려진다 — 오탐")]),
    (F08, [("이OO", "MASK", "내부 IP·연락처가 가려지고, 답변이 코드를 되돌려주면 유출 의심으로 이어진다")]),
]


def extract(path):
    """프론트엔드 추출기(`frontend/src/lib/spreadsheet.js`)와 같은 규칙."""
    ws = load_workbook(path, data_only=True).worksheets[0]
    lines = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else " ".join(str(c).split()) for c in row]
        if any(c.strip() for c in cells):
            lines.append("\t".join(cells).rstrip())
    cols = max(len(r) for r in ws.iter_rows(values_only=True))
    head = f'[표 첨부] {path.name} · 시트 "{ws.title}" · {len(lines)}행 × {cols}열'
    return head + "\n" + "\n".join(lines)


def payload(path):
    return PROMPTS[path.name] + "\n\n" + extract(path)


def post(text, user_id):
    req = urllib.request.Request(
        f"{SERVER}/messages",
        data=json.dumps({"text": text}).encode(),
        headers={"Content-Type": "application/json", "X-User-Id": str(user_id)},
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        return json.loads(e.read())


def verify():
    print(f"\n{'파일':<34}{'계정':<8}{'기대':<9}{'실제':<9}결과")
    print("─" * 96)
    ok = True
    for path, cases in EXPECTED:
        for i, (who, want, _) in enumerate(cases):
            got = post(payload(path), USERS[who]).get("decision", "?")
            mark = "OK" if got == want else "!! 어긋남"
            ok = ok and got == want
            label = path.name if i == 0 else ""
            print(f"{label:<34}{who + '·' + DEPT[USERS[who]]:<14}{want:<9}{got:<9}{mark}")
    print("─" * 96)
    print("전부 일치" if ok else "어긋난 건이 있다 — 시연 전에 고쳐야 한다")
    return ok


def main():
    longest = max(
        (line for line in extract(F08).split("\n") if any(c in line for c in ";{}()")),
        key=len, default="",
    )
    print(f"\n생성 완료 — {OUT_DIR}")
    for path, cases in EXPECTED:
        print(f"  {path.name:<34}{len(extract(path)):>5}자   {cases[0][2]}")
    print(f"\n08번 되돌림 길이 {len(longest)}자 (문턱 40자) — "
          f"{'유출 의심이 뜬다' if len(longest) >= 40 else '⚠ 40자 미만'}")
    if "--verify" in sys.argv:
        sys.exit(0 if verify() else 1)
    print("\n판정 대조는 --verify")


if __name__ == "__main__":
    main()
