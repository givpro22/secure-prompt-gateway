#!/usr/bin/env python3
"""
시연용 엑셀 생성 — 고객 명단 (표 첨부 기본 장면)

시나리오
    영업 담당자가 분기 VIP 명단을 그대로 붙여 "계약금액 상위부터 정리해 줘"라고 묻는다.
    스프레드시트가 사내에서 가장 많이 새는 그릇이고, 새는 방식이 대개 이렇다 —
    악의도 없고 절차를 어긴 것도 아니다. 표 하나를 통째로 복사했을 뿐이다.

    같은 파일에 개인정보가 없는 집계표를 한 장 더 둔다. 규칙이 무엇을 걸고 무엇을
    안 거는지 한 파일로 보이기 위해서다. 다만 추출기는 **첫 시트만** 읽으므로
    두 번째 시트를 보여주려면 시트를 바꿔 저장해야 한다.

    규칙이 못 보는 것도 함께 드러난다. 최도윤·정하윤은 고객 명단에 없는 이름이라
    가려지지 않는다. 규칙은 명단에 있는 문자열만 본다 — 그 한계를 감추지 않는 편이
    시연에서 더 설득력이 있다.

데이터는 전부 합성이다. 실제 고객·연락처·계약이 아니다.
김서준·박예린은 시드 고객 명단에 있는 이름이고, 최도윤·정하윤은 없는 이름이다.
담당자는 기획서 10.1 원칙에 따라 `이OO` 형식을 쓴다.

실행
    /Users/jeonghangyeol/workspace/.venv/bin/python3 make_customer_xlsx.py
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).parent
OUT_NAME = "시연용_고객명단.xlsx"

# 첫 시트 — 검사 대상. 명단에 있는 이름 둘, 없는 이름 둘.
ROSTER = [
    ["고객명", "담당자", "연락처", "이메일", "계약금액", "비고"],
    ["김서준", "이OO", "010-2841-7734", "seojun.kim@sy-tech.co.kr", 120000000, "재계약 협의 중"],
    ["박예린", "김OO", "010-9925-1180", "yerin.park@kkenergy.kr", 84500000, "견적 재요청"],
    ["최도윤", "이OO", "010-3317-4402", "doyun.choi@daehan-log.co.kr", 61200000, "신규 문의"],
    ["정하윤", "김OO", "010-4408-2251", "hayun.jung@nseng.co.kr", 45800000, "보류"],
]

# 둘째 시트 — 개인정보가 없다. 걸릴 것이 없어 그대로 나간다.
SUMMARY = [
    ["월", "신규 상담", "견적 발송", "계약 성사", "성사율(%)"],
    ["7월", 42, 28, 9, 32.1],
    ["8월", 51, 35, 13, 37.1],
    ["9월", 47, 31, 11, 35.5],
]


def style(ws):
    head = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="2F5D8A")
    for cell in ws[1]:
        cell.font, cell.fill = head, fill
        cell.alignment = Alignment(horizontal="center")
    for i, col in enumerate(ws.columns, 1):
        width = max(len(str(c.value or "")) for c in col) + 4
        ws.column_dimensions[get_column_letter(i)].width = max(10, width)
    ws.freeze_panes = "A2"


def write_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "3분기 VIP"
    for row in ROSTER:
        ws.append(row)
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = "#,##0"

    ws2 = wb.create_sheet("월별 집계")
    for row in SUMMARY:
        ws2.append(row)

    style(ws)
    style(ws2)
    wb.save(path)
    return path


def extract_text(path):
    """프론트엔드 추출기(`lib/spreadsheet.js`)와 같다 — 첫 시트만, 탭으로 잇는다."""
    ws = load_workbook(path, data_only=True).worksheets[0]
    lines = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else " ".join(str(c).split()) for c in row]
        if any(c.strip() for c in cells):
            lines.append("\t".join(cells).rstrip())
    return "\n".join(lines)


def main():
    path = write_workbook(OUT_DIR / OUT_NAME)
    text = extract_text(path)

    print(f"\n── 고객 명단 — 표 첨부 기본 장면")
    print(f"   파일          {path.name}")
    print(f"   파일 크기     {path.stat().st_size / 1024:,.1f} KB")
    print(f"   추출 텍스트   {len(text):,} 자 / {text.count(chr(10)) + 1} 행 (첫 시트만)")

    print("\n   가려질 것")
    for label, needles in (
        ("고객명 (명단에 있음)", ("김서준", "박예린")),
        ("연락처", ("010-",)),
        ("이메일", ("@",)),
    ):
        n = sum(text.count(x) for x in needles)
        print(f"     {label:<20} {n}회")

    print("\n   안 가려질 것 — 규칙의 한계를 그대로 보여준다")
    for name in ("최도윤", "정하윤"):
        print(f"     {name} (명단에 없는 이름)   {text.count(name)}회")
    amounts = sum(1 for line in text.split("\n") if any(str(r[4]) in line for r in ROSTER[1:]))
    print(f"     계약금액 (규칙 없음)         {amounts}회")
    print()


if __name__ == "__main__":
    main()
