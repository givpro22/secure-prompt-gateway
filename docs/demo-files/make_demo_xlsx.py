#!/usr/bin/env python3
"""
시연용 엑셀 생성 — 홍보팀 엠바고 데모 (Case E)

시나리오
    개발팀이 4분기 릴리스 백로그를 LLM에 넣어 "스프린트 계획 정리해줘"라고 한다.
    유출 의도도 없고 개인정보도 없다. 그냥 일하려던 것이다.
    그런데 백로그에 홍보팀이 엠바고를 건 신제품명과 런칭 일정이 박혀 있다.
    외부 LLM에 넣는 순간 그건 공개다. 홍보팀 정책이 그것을 막는다.

    같은 파일에 이미 발표된 제품(아틀라스)도 들어 있다. 해제일 2026-09-04가 지났으므로
    걸리지 않는다. 시간 축으로 정책이 갈리는 것이 이 데모의 두 번째 장면이다.

    엠바고 해제일
        SKALA NOVA   2026-09-20  → 발표 당일(09-04) 기준 아직 유효 → 차단
        SKALA ATLAS  2026-09-04  → 그 날부터 공개 가능 → 통과

데이터는 전부 합성이다. 실제 제품명·인명·일정이 아니다.
담당자는 기획서 10.1 원칙에 따라 `이OO` 형식을 쓴다.

결정론적이다. seed를 고정하므로 몇 번을 돌려도 같은 파일이 나온다 —
발표 현장에서 재생성해도 대본의 숫자가 어긋나지 않는다.

실행
    /Users/jeonghangyeol/workspace/.venv/bin/python3 make_demo_xlsx.py
"""

import random
from pathlib import Path

from openpyxl import Workbook

SEED = 20260902
OUT_DIR = Path(__file__).parent

HEADERS = ["항목ID", "에픽", "기능명", "담당", "상태", "스프린트", "비고"]

# --- 엠바고 유효 (해제일 2026-09-20. 발표 당일 기준 아직 못 연다) ------------
NOVA_EPIC = "SKALA NOVA"
NOVA_FEATURES = [
    ("노바 온보딩 플로우 1차", "9/20 런칭 전 필수"),
    ("NOVA 추천 위젯 A/B 테스트", "런칭 2주 전 지표 확정"),
    ("노바 랜딩 페이지 API 연동", "홍보팀 페이지와 동시 오픈"),
    ("SKALA NOVA 결제 모듈 연동", "PG 계약 완료 후 착수"),
    ("노바 베타 피드백 수집 배치", "사내 베타 300명 대상"),
    ("노바 사용량 집계 파이프라인", "9/20 트래픽 대비"),
    ("NOVA 알림 템플릿 다국어화", "KO/EN 우선"),
    ("노바 권한 모델 리팩터링", "정식 오픈 전 마감"),
]

# --- 엠바고 해제 (해제일 2026-09-04. 발표 당일에 풀린다) --------------------
ATLAS_EPIC = "SKALA ATLAS"
ATLAS_FEATURES = [
    ("아틀라스 대시보드 위젯 추가", "정식 출시 후 개선건"),
    ("ATLAS 리포트 PDF 내보내기", "고객 요청 반영"),
    ("아틀라스 필터 성능 개선", "P95 1.2s → 400ms"),
    ("ATLAS 권한 그룹 UI 정리", "운영 이관 완료"),
]

# --- 평범한 개발 항목 (아무 규칙에도 안 걸린다) -----------------------------
PLAIN_EPICS = ["인증·계정", "결제", "검색", "알림", "인프라", "데이터 파이프라인", "관리자 도구"]
PLAIN_FEATURES = [
    ("로그인 세션 만료 처리 개선", "리프레시 경계 조건"),
    ("OAuth 리프레시 토큰 회전", "보안팀 권고 반영"),
    ("결제 실패 재시도 큐 분리", "DLQ 도입"),
    ("상품 검색 색인 재구축 배치", "야간 실행으로 이관"),
    ("푸시 발송 배치 워커 분리", "피크 시 지연 해소"),
    ("Redis 캐시 키 네임스페이스 정리", "충돌 3건 확인"),
    ("CI 파이프라인 의존성 캐시", "빌드 8분 → 3분"),
    ("감사 로그 보관 주기 조정", "180일 → 365일"),
    ("관리자 목록 페이지네이션", "커서 방식 전환"),
    ("이미지 리사이즈 워커 교체", "메모리 누수 해결"),
    ("배치 실패 알림 채널 정리", "중복 알림 억제"),
    ("DB 커넥션 풀 상한 조정", "부하 테스트 결과 반영"),
    ("정적 자원 CDN 캐시 헤더", "TTL 재설정"),
    ("에러 코드 체계 통일", "문서화 병행"),
    ("헬스체크 엔드포인트 분리", "readiness/liveness"),
    ("사용자 설정 마이그레이션", "레거시 컬럼 제거"),
    ("검색 자동완성 응답 축소", "페이로드 40% 감소"),
    ("주문 상태 전이 검증 추가", "불일치 케이스 차단"),
    ("리포트 쿼리 인덱스 추가", "풀스캔 제거"),
    ("테스트 픽스처 정리", "중복 제거"),
]

OWNERS = ["이OO", "김OO", "정OO", "박OO", "최OO"]
STATUSES = ["개발중", "리뷰", "완료", "대기"]
SPRINTS = ["S-22", "S-23", "S-24", "S-25", "S-26"]


def build_rows(total, nova_ratio, atlas_ratio, rnd):
    """백로그 행을 조립한다. 엠바고 대상 행이 파일 전반에 흩어지게 섞는다."""
    nova_n = int(total * nova_ratio)
    atlas_n = int(total * atlas_ratio)
    plain_n = total - nova_n - atlas_n

    rows = []
    for i in range(nova_n):
        name, memo = NOVA_FEATURES[i % len(NOVA_FEATURES)]
        rows.append((NOVA_EPIC, name, memo))
    for i in range(atlas_n):
        name, memo = ATLAS_FEATURES[i % len(ATLAS_FEATURES)]
        rows.append((ATLAS_EPIC, name, memo))
    for i in range(plain_n):
        name, memo = PLAIN_FEATURES[i % len(PLAIN_FEATURES)]
        rows.append((rnd.choice(PLAIN_EPICS), name, memo))

    rnd.shuffle(rows)

    out = []
    for idx, (epic, name, memo) in enumerate(rows, start=1):
        out.append([
            f"REL-{idx:04d}",
            epic,
            name,
            OWNERS[idx % len(OWNERS)],
            STATUSES[idx % len(STATUSES)],
            SPRINTS[idx % len(SPRINTS)],
            memo,
        ])
    return out


def write_workbook(path, rows, hidden_schedule):
    wb = Workbook()
    ws = wb.active
    ws.title = "백로그"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    for col, width in zip("ABCDEFG", (10, 14, 32, 8, 8, 9, 26)):
        ws.column_dimensions[col].width = width

    if hidden_schedule:
        # 숨긴 시트. 본문 셀만 검사하면 놓치는 자리다 — 추출기가 숨긴 시트까지
        # 읽는다는 것을 시연에서 보여주는 용도다.
        ws2 = wb.create_sheet("런칭_일정")
        ws2.append(["제품", "코드명", "대외 발표일", "엠바고 해제", "관리 부서"])
        ws2.append(["SKALA NOVA", "노바", "2026-09-20", "2026-09-20", "홍보팀"])
        ws2.append(["SKALA ATLAS", "아틀라스", "2026-09-04", "2026-09-04", "홍보팀"])
        ws2.append(["SKALA CORE", "코어", "2026-03-02", "2026-03-02", "홍보팀"])
        for col, width in zip("ABCDE", (16, 10, 14, 14, 10)):
            ws2.column_dimensions[col].width = width
        ws2.sheet_state = "hidden"

    wb.save(path)
    return path


def extract_text(path):
    """프론트엔드 추출기가 만들 텍스트를 그대로 흉내 낸다.

    `[시트명!N행] 셀 | 셀 | ...` 형식이라, 오프셋 기반 마스킹이 그대로 동작하면서
    차단 사유에 위치가 공짜로 딸려 나온다. 숨긴 시트도 포함한다.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    lines = []
    for ws in wb.worksheets:
        for n, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = [str(c) for c in row if c is not None]
            if cells:
                lines.append(f"[{ws.title}!{n}행] " + " | ".join(cells))
    return "\n".join(lines)


def count_keyword(text, keyword):
    return text.count(keyword)


def report(label, path, text):
    size_kb = path.stat().st_size / 1024
    print(f"\n── {label}")
    print(f"   파일          {path.name}")
    print(f"   파일 크기     {size_kb:,.1f} KB")
    print(f"   추출 텍스트   {len(text):,} 자")
    print(f"   추출 행 수    {text.count(chr(10)) + 1:,} 행")
    for kw in ("노바", "NOVA", "아틀라스", "ATLAS", "9/20", "2026-09-20", "2026-09-04"):
        c = count_keyword(text, kw)
        if c:
            print(f"   '{kw}' 등장   {c:,} 회")


def main():
    rnd = random.Random(SEED)

    # 파일 A — 시연 본편. 엠바고에 걸려 차단된다.
    rows_a = build_rows(total=220, nova_ratio=0.30, atlas_ratio=0.10, rnd=rnd)
    path_a = write_workbook(OUT_DIR / "2026_4Q_릴리스_백로그.xlsx", rows_a, hidden_schedule=True)
    text_a = extract_text(path_a)
    report("파일 A — 검사 대상 (엠바고 차단)", path_a, text_a)

    # 파일 B — 상한 초과. 검사 전에 거절된다.
    rnd_b = random.Random(SEED + 1)
    rows_b = build_rows(total=2600, nova_ratio=0.20, atlas_ratio=0.10, rnd=rnd_b)
    path_b = write_workbook(OUT_DIR / "전체_제품_백로그_아카이브.xlsx", rows_b, hidden_schedule=False)
    text_b = extract_text(path_b)
    report("파일 B — 상한 초과 (검사 전 거절)", path_b, text_b)

    print("\n── 상한 제안")
    print(f"   추출 텍스트 상한   50,000 자   (파일 A의 약 {50000 / len(text_a):.1f}배, 파일 B는 초과)")
    print(f"   파일 크기 상한     2 MB       (파일 A의 약 {2 * 1024 * 1024 / path_a.stat().st_size:.0f}배)")
    print()


if __name__ == "__main__":
    main()
