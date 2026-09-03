#!/usr/bin/env python3
"""
시연용 엑셀 생성 — 장애 리포트 (표 첨부 + 원문 유출 검토를 한 번에)

시나리오
    개발팀이 결제 재시도 NPE 장애를 정리한 리포트를 받았다. 사내에서 늘 도는
    형태의 표다 — 티켓 번호, 담당자 연락처, DB 호스트, 영향 고객, 그리고 재현 코드.
    그대로 외부 모델에 붙여 "원인 좀 봐 줘"라고 묻는다.

    한 파일에서 두 장면이 이어진다.

    1) 나갈 때 — 표에서 뽑은 텍스트가 정책 검사를 거친다.
       내부 IP·연락처·고객명이 라벨로 바뀌고 마스킹본만 나간다.

    2) 돌아올 때 — 모델은 고쳐 준 코드를 답변에 통째로 되돌려준다.
       사내 코드가 외부를 한 바퀴 돌아 그대로 나온 셈이라 유출 검사가 걸고,
       보안 담당자가 확정한다.

    이 두 번째 장면이 성립하려면 되돌아온 조각이 40자를 넘어야 한다. 스크립트가
    끝에 그 길이를 재서 알려 준다 — 셀 내용을 손보다가 문턱 아래로 떨어지면
    시연에서 유출 의심이 안 뜬다.

데이터는 전부 합성이다. 실제 티켓·인명·주소가 아니다.
담당자는 기획서 10.1 원칙에 따라 `이OO` 형식을 쓴다.
고객명은 시드에 있는 이름을 쓴다 — 명단 규칙이 걸리는 것을 보여야 하기 때문이다.

실행
    /Users/jeonghangyeol/workspace/.venv/bin/python3 make_incident_xlsx.py
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

OUT_DIR = Path(__file__).parent
OUT_NAME = "장애리포트_결제재시도.xlsx"

# 재현 코드. 한 셀에 넣는다 — 추출기가 셀 안의 줄바꿈을 공백으로 바꿔 한 줄로 만든다.
# 그 한 줄이 그대로 답변에 되돌아오는 것이 유출 검사가 잡는 장면이다.
REPRO_CODE = (
    'RetryPolicy policy = config.getRetryPolicy();\n'
    'int max = policy.getMaxAttempts();\n'
    'client.connect("10.0.3.21", 5432);'
)

ROWS = [
    ["항목", "내용"],
    ["티켓 번호", "INC-2026-0903-17"],
    ["발생 시각", "2026-09-03 14:22"],
    ["발생 모듈", "결제 재시도 (payment-retry)"],
    ["심각도", "P2"],
    ["증상", "결제 재시도에서 NPE 나는데 봐줘."],
    ["재현 코드", REPRO_CODE],
    ["DB 호스트", "10.0.3.21:5432"],
    ["담당자", "이OO"],
    ["담당자 연락처", "010-3317-4402"],
    ["영향 고객", "김서준"],
    ["고객 연락처", "010-2841-7734"],
    ["조치 상태", "원인 분석 중"],
]


def write_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "장애 리포트"
    for row in ROWS:
        ws.append(row)

    head = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="2F5D8A")
    for cell in ws[1]:
        cell.font, cell.fill = head, fill
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 62
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[7].height = 48
    ws.freeze_panes = "A2"

    wb.save(path)
    return path


def extract_text(path):
    """프론트엔드 추출기(`lib/spreadsheet.js`)가 만들 텍스트를 그대로 흉내 낸다.

    첫 시트만, 빈 행은 빼고, 셀은 탭으로 잇는다. 셀 안의 줄바꿈은 공백이 된다.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    lines = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else " ".join(str(c).split()) for c in row]
        if any(c.strip() for c in cells):
            lines.append("\t".join(cells).rstrip())
    return "\n".join(lines)


def longest_code_run(text):
    """`;{}()` 가 든 줄 중 가장 긴 것. 유출 검사(40자)가 무엇을 잡을지 미리 본다."""
    runs = [line for line in text.split("\n") if any(ch in line for ch in ";{}()")]
    return max(runs, key=len) if runs else ""


def main():
    path = write_workbook(OUT_DIR / OUT_NAME)
    text = extract_text(path)
    run = longest_code_run(text)

    print(f"\n── 장애 리포트 — 표 첨부 + 원문 유출 검토")
    print(f"   파일          {path.name}")
    print(f"   파일 크기     {path.stat().st_size / 1024:,.1f} KB")
    print(f"   추출 텍스트   {len(text):,} 자 / {text.count(chr(10)) + 1} 행")

    print("\n   나갈 때 걸릴 것 (규칙이 라벨로 바꾼다)")
    for label, needle in (
        ("내부 IP", "10.0.3.21"),
        ("담당자 연락처", "010-3317-4402"),
        ("고객 연락처", "010-2841-7734"),
        ("고객명", "김서준"),
    ):
        hits = text.count(needle)
        print(f"     {label:<14} {needle:<16} {hits}회")

    print("\n   돌아올 때 걸릴 것 (답변이 되돌려주면 유출 의심)")
    print(f"     가장 긴 코드 줄  {len(run)}자  (문턱 40자)")
    print(f"     {'통과 — 유출 의심이 뜬다' if len(run) >= 40 else '⚠ 40자 미만 — 유출 의심이 안 뜬다'}")
    print(f"     {run[:90]}…")
    print()


if __name__ == "__main__":
    main()
